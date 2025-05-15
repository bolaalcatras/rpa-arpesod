import imaplib
import email
from email.header import decode_header
import openpyxl
from datetime import datetime
import time
import traceback

# === CONFIGURACIÓN ===
EMAIL = "activosfijoselectrocreditosn@gmail.com"
PASSWORD = "ddqy vvzl ziqb wzpp"
REMITE = "activosfijoselectrocreditosn@gmail.com"
RUTA_EXCEL = "C:/Users/usuario/Desktop/PASANTE SENA SISTEMAS/Indicador.xlsx"
LOG_PATH = "C:/Users/usuario/Desktop/PASANTE SENA SISTEMAS/log_correo.txt"
INTERVALO_SEGUNDOS = 60  # tiempo entre revisiones
ultimo_id_visto = None   # para evitar registrar el mismo correo varias veces

# === BUCLE INFINITO ===
while True:
    try:
        mail = imaplib.IMAP4_SSL("imap.gmail.com")
        mail.login(EMAIL, PASSWORD)
        mail.select("inbox")

        status, messages = mail.search(None, f'FROM "{REMITE}"')
        email_ids = messages[0].split()
    
        if email_ids:
            latest_email_id = email_ids[-1]

            if latest_email_id != ultimo_id_visto:
                status, msg_data = mail.fetch(latest_email_id, "(RFC822)")
                msg = email.message_from_bytes(msg_data[0][1])

                # Abrimos el Excel y escribimos
                wb = openpyxl.load_workbook(RUTA_EXCEL)
                ws = wb.active
                now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                ws['D4'] = "Correo recibido"
                ws['C4'] = now
                wb.save(RUTA_EXCEL)

                # Guardamos log
                with open(LOG_PATH, "a", encoding="utf-8") as log:
                    log.write(f"[{now}] Correo detectado y Excel actualizado.\n")

                print("✔ Correo detectado y registrado.")
                ultimo_id_visto = latest_email_id
            else:
                print("⏳ Mismo correo, esperando nuevo.")
        else:
            print("⚠ No hay correos nuevos del remitente.")

        mail.logout()

    except Exception as e:
        with open(LOG_PATH, "a", encoding="utf-8") as log:
            log.write(f"[{datetime.now()}] ERROR: {str(e)}\n{traceback.format_exc()}\n")
        print("❌ Error. Ver log.")

    # Espera antes de volver a revisar
    time.sleep(INTERVALO_SEGUNDOS)