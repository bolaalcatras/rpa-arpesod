import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from copy import copy

# ========== FUNCIONES ==========

def cargar_datos_excel(path, sheet_name):
    df = pd.read_excel(path, sheet_name=sheet_name)
    df = df[df["COD_REGION"].notna() & (df["COD_REGION"] != 999)]
    df = df[df["FORMAPAGO"].notna() & (df["FORMAPAGO"] != 'NO APLICA')]
    df = df.sort_values(by="FECHA_FACT", ascending=False)

    # Convertir la columna a string en el formato deseado
    df["FECHA_FACT"] = df["FECHA_FACT"].dt.strftime("%d-%m-%Y")

    # Obtener la fecha más reciente
    fecha = df["FECHA_FACT"].iloc[0]

    return df, fecha

def crear_pivot(df):
    pivot = pd.pivot_table(df, index="COD_REGION", columns="FORMAPAGO", values="VTAS_ANT_I", aggfunc="sum", fill_value=0)
    pivot2 = pd.pivot_table(df, index="COD_REGION", columns="DNONOMBRE", values="VTAS_ANT_I", aggfunc="sum", fill_value=0)
    return pivot, pivot2

def filtrar_datos1(pivot, pivot2):
    return {
        'CONTADO': pivot[['CONTADO']],
        'CREDICONTADO': pivot[['CREDICONTADO']],
        'CREDITO': pivot[['CREDITO']],
        'VENTAS BRILLA': pivot2[['VENTAS BRILLA']],
        'VENTAS SISTECREDITO': pivot2[['VENTAS SISTECREDITO']],
        'VENTAS ADDI': pivot2[['VENTAS ADDI']],
    }

def filtrar_datos2(pivot3, pivot4):
    return {
        'CONTADO': pivot3[['CONTADO']],
        'CREDICONTADO': pivot3[['CREDICONTADO']],
        'CREDITO': pivot3[['CREDITO']],
        'VENTAS BRILLA': pivot4[['VENTAS BRILLA']],
        'VENTAS SISTECREDITO': pivot4[['VENTAS SISTECREDITO']],
        'VENTAS ADDI': pivot4[['VENTAS ADDI']],
    }
    
def filtrar_datos3(pivot5, pivot6):
    return {
        'CONTADO': pivot5[['CONTADO']],
        'CREDICONTADO': pivot5[['CREDICONTADO']],
        'CREDITO': pivot5[['CREDITO']],
        'VENTAS BRILLA': pivot6[['VENTAS BRILLA']],
        'VENTAS SISTECREDITO': pivot6[['VENTAS SISTECREDITO']],
    }        

def cargar_o_crear_libro(path):
    try:
        wb = load_workbook(path)
    except FileNotFoundError:
        wb = Workbook()
        wb.remove(wb.active)
    return wb

def crear_hoja_si_no_existe(wb, nombre_hoja):
    if nombre_hoja in wb.sheetnames:
        return wb[nombre_hoja]
    else:
        return wb.create_sheet(title=nombre_hoja)

def copiar_hoja(ws_origen, ws_dest):
    for row in ws_origen.iter_rows():
        for cell in row:
            new_cell = ws_dest.cell(row=cell.row, column=cell.col_idx, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    for col in ws_origen.column_dimensions:
        ws_dest.column_dimensions[col].width = ws_origen.column_dimensions[col].width
        ws_dest.column_dimensions[col].hidden = ws_origen.column_dimensions[col].hidden

    for row_dim in ws_origen.row_dimensions:
        ws_dest.row_dimensions[row_dim].height = ws_origen.row_dimensions[row_dim].height
        ws_dest.row_dimensions[row_dim].hidden = ws_origen.row_dimensions[row_dim].hidden

def insertar_valores1(ws_dest, datos):
    for i, valor in enumerate(datos['CONTADO']['CONTADO'], start=4):
        ws_dest.cell(row=i, column=4, value=valor)
    for i, valor in enumerate(datos['CREDICONTADO']['CREDICONTADO'], start=4):
        ws_dest.cell(row=i, column=8, value=valor)
    for i, valor in enumerate(datos['CREDITO']['CREDITO'], start=4):
        ws_dest.cell(row=i, column=12, value=valor)
    for i, valor in enumerate(datos['VENTAS BRILLA']['VENTAS BRILLA'], start=4):
        ws_dest.cell(row=i, column=16, value=valor)
    for i, valor in enumerate(datos['VENTAS SISTECREDITO']['VENTAS SISTECREDITO'], start=4):
        ws_dest.cell(row=i, column=17, value=valor)
    for i, valor in enumerate(datos['VENTAS ADDI']['VENTAS ADDI'], start=4):
        ws_dest.cell(row=i, column=18, value=valor)
        
def insertar_valores2(ws_dest, datos):
    for i, valor in enumerate(datos['CONTADO']['CONTADO'], start=17):
        ws_dest.cell(row=i, column=4, value=valor)
    for i, valor in enumerate(datos['CREDICONTADO']['CREDICONTADO'], start=17):
        ws_dest.cell(row=i, column=8, value=valor)
    for i, valor in enumerate(datos['CREDITO']['CREDITO'], start=17):
        ws_dest.cell(row=i, column=12, value=valor)
    for i, valor in enumerate(datos['VENTAS BRILLA']['VENTAS BRILLA'], start=17):
        ws_dest.cell(row=i, column=16, value=valor)
    for i, valor in enumerate(datos['VENTAS SISTECREDITO']['VENTAS SISTECREDITO'], start=17):
        ws_dest.cell(row=i, column=17, value=valor)
    for i, valor in enumerate(datos['VENTAS ADDI']['VENTAS ADDI'], start=17):
        ws_dest.cell(row=i, column=18, value=valor)

def insertar_valores3(ws_dest, datos):
    for i, valor in enumerate(datos['CONTADO']['CONTADO'], start=30):
        ws_dest.cell(row=i, column=4, value=valor)
    for i, valor in enumerate(datos['CREDICONTADO']['CREDICONTADO'], start=30):
        ws_dest.cell(row=i, column=8, value=valor)
    for i, valor in enumerate(datos['CREDITO']['CREDITO'], start=30):
        ws_dest.cell(row=i, column=12, value=valor)
    for i, valor in enumerate(datos['VENTAS BRILLA']['VENTAS BRILLA'], start=30):
        ws_dest.cell(row=i, column=16, value=valor)
    for i, valor in enumerate(datos['VENTAS SISTECREDITO']['VENTAS SISTECREDITO'], start=30):
        ws_dest.cell(row=i, column=17, value=valor)


# ========== PROCESO PRINCIPAL ==========

def procesar_archivo():
    archivo1 = "/home/alexander/Descargas/CRTMPCONSULTA1.XLSX"
    archivo2 = "/home/alexander/Descargas/CRTMPCONSULTA 2.XLSX"
    archivo3 = "/home/alexander/Descargas/CRTMPCONSULTA 3.XLSX"
    nuevo_archivo = "/home/alexander/Descargas/VentasDiarias2025 MAYO .xlsx"
    hoja_origen = '13-05-2025'

    df1, fecha_actual = cargar_datos_excel(archivo1, 'Page 001')
    df2, fecha_actual = cargar_datos_excel(archivo2, 'Page 001')
    df3, fecha_actual = cargar_datos_excel(archivo3, 'Page 001')
    
    pivot, pivot2 = crear_pivot(df1)
    pivot3, pivot4 = crear_pivot(df2)
    pivot5, pivot6 = crear_pivot(df3)
    
    datos_filtrados = filtrar_datos1(pivot, pivot2)
    datos_filtrados2 = filtrar_datos2(pivot3, pivot4)
    datos_filtrados3 = filtrar_datos3(pivot5, pivot6)

    wb = cargar_o_crear_libro(nuevo_archivo)
    ws_dest = crear_hoja_si_no_existe(wb, fecha_actual)
    ws_origen = wb[hoja_origen]

    copiar_hoja(ws_origen, ws_dest)
    insertar_valores1(ws_dest, datos_filtrados)
    insertar_valores2(ws_dest, datos_filtrados2)
    insertar_valores3(ws_dest, datos_filtrados3)

    wb.save(nuevo_archivo)
    print(f"Contenido copiado de '{hoja_origen}' a '{fecha_actual}' y guardado en {nuevo_archivo}")

# Ejecutar
if __name__ == "__main__":
    procesar_archivo()